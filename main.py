import datetime

import openpyxl

import memrise

import time

import os


__location__ = os.path.realpath(os.path.join(os.getcwd(), os.path.dirname(__file__)))

the_time = datetime.datetime.now()

dateis = the_time.strftime('%d-%m-%Y')

wb = openpyxl.load_workbook('eg.xlsx')
my_req_sheet = wb.get_sheet_by_name('Vocab')
time.sleep(5)

# to read number from file
filer = open(os.path.join(__location__, 'pipe.txt'))
#filer = open('D:/Memrise vocab tracker/pipe.txt','r')
NUM = filer.read()


list1 = [23]


while (my_req_sheet.cell(row =list1[0] , column= 4).value is not None):
    list1[0] = list1[0] + 1


my_req_sheet.cell(row=list1[0], column=4).value = int(NUM)
my_req_sheet.cell(row=list1[0], column=3).value = dateis

wb.save('eg.xlsx')
